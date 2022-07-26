import pathlib
from openpyxl import Workbook, load_workbook
# https://openpyxl.readthedocs.io/en/stable/

root = pathlib.Path(__file__).parent
files_path = root.joinpath("files")

# read xlsx files
wb = load_workbook(files_path.joinpath("inventory.xlsx"))
print(wb.sheetnames)
sheet1 = wb['Sheet']

print(sheet1['A1'].value)



# write xlsx files
filename = "inventory_out.xlsx"

# init workbook
workbook = Workbook()
sheet = workbook.active

# header - how to highlight
sheet["A1"] = "Produs"
sheet["B1"] = "U.M."
sheet["C1"] = "Cantitate"

sheet["A2"] = "Vopsea"
sheet["B2"] = "BUC"
sheet["C2"] = 10


# workbook.save(filename=files_path.joinpath(filename))