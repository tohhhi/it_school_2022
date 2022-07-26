import csv
import pathlib
import pathlib
from openpyxl import Workbook, load_workbook
import pandas as pd

root = pathlib.Path(__file__).parent
files_path = root.joinpath("files")



# try:
#     with open(files_path.joinpath("salarii.csv")) as fin:
#         reader = list(csv.reader(fin))
# except OSError:
#     print("File error.")
# else:
#     lista_salarii = []
#     for i in reader:
        
#         print(i)
#         lista_salarii.append(float(i[3]))

#     print(f"Total salarii: {sum(lista_salarii)}")  


# citire fisier csv cu nume campuri
field_names = [
    "first_name",
    "last_name",
    "id",
    "gros_salary",
    "days_off"
]
try:
    with open(files_path.joinpath("salarii.csv")) as fin:
        # list -> extrage datele din generator
        dict_reader = list(csv.DictReader(fin, fieldnames=field_names))
        for i in dict_reader:
            print(i.values())
except OSError:
    print("File error.")





root = pathlib.Path(__file__).parent
files_path = root.joinpath("files")

# write xlsx files
filename = "convert_csv_to_xlsx.xlsx"

# init workbook
workbook = Workbook()
sheet = workbook.active

# header - how to highlight
# sheet["A1"] = "Nume"
# sheet["B1"] = "Prenume"
# sheet["C1"] = "Id"


# sheet["A2"] = i["first_name"]
# sheet["B2"] = i["last_name"]
# sheet["C2"] = i["id"]


list_sheet = workbook.create_sheet('list')

# Reading the csv file
df_new = pd.read_csv('salarii.csv')
 
# saving xlsx file
GFG = pd.ExcelWriter('v2Names.xlsx')
df_new.to_excel(GFG, index=False)
 
GFG.save()



# workbook.save(filename=files_path.joinpath(filename))